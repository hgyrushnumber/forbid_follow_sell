#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import time
from typing import List, Optional
import datetime
from datetime import datetime as dt, timezone
from playwright.sync_api import TimeoutError as PlaywrightTimeoutError

# 配置常量
DEFAULT_STORAGE_DIR = "accounts"
DEBUG_DIR = "debug_artifacts"

# MENU_BUTTONS常量已移至services.constants模块

USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/122.0.0.0 Safari/537.36"
)

CANDIDATE_HEADERS = [
    "sku",
    "seller sku",
    "seller_sku",
    "sku id",
    "商品sku",
    "商品 sku",
    "商家sku",
    "商家 sku",
    "货号",
    "编码",
    "артикул",
    "sku продавца",
]

_LOGGER = print


def set_logger(logger_func):
    global _LOGGER
    _LOGGER = logger_func


def log(msg: str):
    _LOGGER(msg)


def sleep(ms: int):
    time.sleep(ms / 1000)


def ensure_dirs():
    os.makedirs(DEFAULT_STORAGE_DIR, exist_ok=True)
    os.makedirs(DEBUG_DIR, exist_ok=True)


def debug_path(name: str, ext: str) -> str:
    return os.path.join(DEBUG_DIR, f"{name}.{ext}")


def save_login_state(context, path: str):
    try:
        context.storage_state(path=path)
        log(f"✅ 登录态已导出到 {path}")
        return True
    except Exception as e:
        log(f"⚠️ 导出登录态失败: {e}")
        return False


def print_page_debug(page, tag="DEBUG"):
    try:
        log(f"[{tag}] URL: {page.url}")
    except Exception:
        pass

    try:
        log(f"[{tag}] Title: {page.title()}")
    except Exception:
        pass

    try:
        body_text = page.locator("body").inner_text(timeout=5000)
        log(f"[{tag}] 页面文本前1000字符:\n{body_text[:1000]}\n")
    except Exception as e:
        log(f"[{tag}] 读取 body 文本失败: {e}")


def dump_page_state(page, tag):
    log(f"\n===== DUMP PAGE STATE: {tag} =====")
    print_page_debug(page, tag)

    try:
        html = page.content()
        html_file = debug_path(tag, "html")
        with open(html_file, "w", encoding="utf-8") as f:
            f.write(html)
        log(f"📝 已保存 HTML: {html_file}")
    except Exception as e:
        log(f"⚠️ 保存 HTML 失败: {e}")

    try:
        png_file = debug_path(tag, "png")
        page.screenshot(path=png_file, full_page=True)
        log(f"📸 已保存截图: {png_file}")
    except Exception as e:
        log(f"⚠️ 保存截图失败: {e}")


def safe_body_text(page, timeout=3000):
    try:
        return page.locator("body").inner_text(timeout=timeout)
    except Exception:
        return ""


def normalize_header(v) -> str:
    return str(v or "").strip().lower()


def read_skus_from_excel(excel_path: str) -> List[str]:
    from openpyxl import load_workbook
    wb = load_workbook(excel_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise ValueError("Excel 数据为空或只有表头")

    header_row = [normalize_header(v) for v in rows[0]]

    sku_col_index = -1
    for i, cell in enumerate(header_row):
        if not cell:
            continue
        if cell in CANDIDATE_HEADERS or "sku" in cell or "артикул" in cell:
            sku_col_index = i
            break

    if sku_col_index == -1:
        raise ValueError("未找到 SKU 列，请检查表头是否包含 SKU / 商家SKU / Артикул")

    skus = []
    for row in rows[1:]:
        if row is None or sku_col_index >= len(row):
            continue
        value = row[sku_col_index]
        if value is None:
            continue
        sku = str(value).strip()
        if sku:
            skus.append(sku)

    return list(dict.fromkeys(skus))


def escape_xpath_text(text: str) -> str:
    if "'" not in text:
        return f"'{text}'"
    if '"' not in text:
        return f'"{text}"'
    parts = text.split("'")
    return "concat(" + ", '\'', ".join(f"'{p}'" for p in parts) + ")"


def build_text_xpaths(text: str, ru_text: Optional[str] = None) -> List[str]:
    values = []
    if text:
        values.append(text)
    if ru_text:
        values.append(ru_text)

    selectors = []
    for value in values:
        safe = escape_xpath_text(value)
        selectors.extend([
            f"//span[contains(normalize-space(.), {safe})]",
            f"//span[normalize-space(text())={safe}]",
            f"//button[contains(normalize-space(.), {safe})]",
            f"//button[normalize-space(text())={safe}]",
            f"//a[contains(normalize-space(.), {safe})]",
            f"//a[normalize-space(text())={safe}]",
            f"//*[contains(normalize-space(.), {safe})]",
        ])
    return selectors


def find_visible_by_xpaths(page, selectors: List[str], timeout_ms: int = 4000):
    deadline = time.time() + timeout_ms / 1000
    while time.time() < deadline:
        for xp in selectors:
            locator = page.locator(f"xpath={xp}")
            try:
                if locator.count() > 0 and locator.first.is_visible():
                    return locator.first
            except Exception:
                pass
        sleep(300)
    return None


# IMAP相关功能已全部删除，如需恢复请查看历史版本