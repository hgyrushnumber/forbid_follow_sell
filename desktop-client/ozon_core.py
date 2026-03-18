import os
import time
from datetime import datetime, timezone
from typing import List, Optional, Callable, Dict, Any
from dataclasses import dataclass, field

from openpyxl import load_workbook
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError
from playwright_stealth import Stealth

from email_otp import get_otp_from_email, get_latest_ozon_mail_id
import sys
from models import OzonAccount, BrowserSession
if getattr(sys, "frozen", False):
    os.environ["PLAYWRIGHT_BROWSERS_PATH"] = os.path.join(sys._MEIPASS, "ms-playwright")

# =========================
# 配置常量
# =========================
HEADLESS = False
SLOW_MO = 200

TARGET_URL = "https://seller.ozon.ru/app/messenger?channel=SCRM"
DASHBOARD_URL = "https://seller.ozon.ru/app/dashboard/main"
HOME_URL = "https://seller.ozon.ru/"

from services.utils import ensure_dirs, log, sleep, set_logger
from services.session_service import SessionService
from services.page_service import PageService
from services.sku_service import SkuService
from services.page_state_detector import detect_page_type as _detect_page_type, is_messenger_page as _is_messenger_page



# =========================
# 全局实例
# =========================
session_service: Optional['SessionService'] = None
page_service: Optional['PageService'] = None
sku_service: Optional['SkuService'] = None
_LOGGER: Callable[[str], None] = print


def set_logger(logger_func: Callable[[str], None]):
    global _LOGGER, session_service, page_service, sku_service
    _LOGGER = logger_func

    # 延迟导入避免循环依赖
    from services.session_service import SessionService
    from services.page_service import PageService
    from services.sku_service import SkuService

    # 初始化服务
    session_service = SessionService(logger_func)
    page_service = PageService(logger_func)
    sku_service = SkuService(logger_func)


def log(msg: str):
    _LOGGER(msg)


def sleep(ms: int):
    time.sleep(ms / 1000)


# =========================
# 通用辅助
# =========================
def ensure_dirs():
    os.makedirs(DEFAULT_STORAGE_DIR, exist_ok=True)
    os.makedirs(VIDEO_DIR, exist_ok=True)
    os.makedirs(DEBUG_DIR, exist_ok=True)


def debug_path(name: str, ext: str) -> str:
    return os.path.join(DEBUG_DIR, f"{name}.{ext}")


def save_login_state(context, path: str):
    try:
        context.storage_state(path=path)
        log(f"✅ 登录态已导出到 {path}")
        return True
    except Exception as e:
        log(f"⚠️ 导出登录态失败: {e}")
        return False


def print_page_debug(page, tag="DEBUG"):
    try:
        log(f"[{tag}] URL: {page.url}")
    except Exception:
        pass

    try:
        log(f"[{tag}] Title: {page.title()}")
    except Exception:
        pass

    try:
        body_text = page.locator("body").inner_text(timeout=5000)
        log(f"[{tag}] 页面文本前1000字符:\n{body_text[:1000]}\n")
    except Exception as e:
        log(f"[{tag}] 读取 body 文本失败: {e}")


def dump_page_state(page, tag):
    log(f"\n===== DUMP PAGE STATE: {tag} =====")
    print_page_debug(page, tag)

    try:
        html = page.content()
        html_file = debug_path(tag, "html")
        with open(html_file, "w", encoding="utf-8") as f:
            f.write(html)
        log(f"📝 已保存 HTML: {html_file}")
    except Exception as e:
        log(f"⚠️ 保存 HTML 失败: {e}")

    try:
        png_file = debug_path(tag, "png")
        page.screenshot(path=png_file, full_page=True)
        log(f"📸 已保存截图: {png_file}")
    except Exception as e:
        log(f"⚠️ 保存截图失败: {e}")


def attach_debug_listeners(page):
    def on_request(request):
        try:
            log(f"[REQUEST] {request.method} {request.url}")
            if request.method in ("POST", "PUT", "PATCH"):
                data = request.post_data
                if data:
                    log(f"[REQUEST BODY] {data[:1500]}")
        except Exception as e:
            log(f"[REQUEST ERROR] {e}")

    def on_response(response):
        try:
            log(f"[RESPONSE] {response.status} {response.url}")
            url = response.url.lower()

            if any(
                k in url
                for k in [
                    "otp",
                    "auth",
                    "verify",
                    "login",
                    "code",
                    "id.ozon",
                    "widget/json/v2",
                    "_action/emailotpentry",
                ]
            ):
                try:
                    text = response.text()
                    log(f"[RESPONSE BODY] {text[:3000]}")
                except Exception:
                    pass
        except Exception as e:
            log(f"[RESPONSE ERROR] {e}")

    def on_request_failed(request):
        try:
            log(f"[REQUEST FAILED] {request.method} {request.url} -> {request.failure}")
        except Exception as e:
            log(f"[REQUEST FAILED ERROR] {e}")

    def on_console(msg):
        try:
            log(f"[CONSOLE] {msg.type}: {msg.text}")
        except Exception:
            pass

    def on_page_error(error):
        log(f"[PAGE ERROR] {error}")

    page.on("request", on_request)
    page.on("response", on_response)
    page.on("requestfailed", on_request_failed)
    page.on("console", on_console)
    page.on("pageerror", on_page_error)


def safe_body_text(page, timeout=3000):
    try:
        return page.locator("body").inner_text(timeout=timeout)
    except Exception:
        return ""


# =========================
# Excel
# =========================
def normalize_header(v) -> str:
    return str(v or "").strip().lower()


def read_skus_from_excel(excel_path: str) -> List[str]:
    wb = load_workbook(excel_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise ValueError("Excel 数据为空或只有表头")

    header_row = [normalize_header(v) for v in rows[0]]

    sku_col_index = -1
    for i, cell in enumerate(header_row):
        if not cell:
            continue
        if cell in CANDIDATE_HEADERS or "sku" in cell or "артикул" in cell:
            sku_col_index = i
            break

    if sku_col_index == -1:
        raise ValueError("未找到 SKU 列，请检查表头是否包含 SKU / 商家SKU / Артикул")

    skus = []
    for row in rows[1:]:
        if row is None or sku_col_index >= len(row):
            continue
        value = row[sku_col_index]
        if value is None:
            continue
        sku = str(value).strip()
        if sku:
            skus.append(sku)

    return list(dict.fromkeys(skus))


# =========================
# XPath / 页面查找
# =========================
def escape_xpath_text(text: str) -> str:
    if "'" not in text:
        return f"'{text}'"
    if '"' not in text:
        return f'"{text}"'
    parts = text.split("'")
    return "concat(" + ", \"'\", ".join(f"'{p}'" for p in parts) + ")"


def build_text_xpaths(text: str, ru_text: Optional[str] = None) -> List[str]:
    values = []
    if text:
        values.append(text)
    if ru_text:
        values.append(ru_text)

    selectors = []
    for value in values:
        safe = escape_xpath_text(value)
        selectors.extend([
            f"//span[contains(normalize-space(.), {safe})]",
            f"//span[normalize-space(text())={safe}]",
            f"//button[contains(normalize-space(.), {safe})]",
            f"//button[normalize-space(text())={safe}]",
            f"//a[contains(normalize-space(.), {safe})]",
            f"//a[normalize-space(text())={safe}]",
            f"//*[contains(normalize-space(.), {safe})]",
        ])
    return selectors


def find_visible_by_xpaths(page, selectors: List[str], timeout_ms: int = 4000):
    deadline = time.time() + timeout_ms / 1000
    while time.time() < deadline:
        for xp in selectors:
            locator = page.locator(f"xpath={xp}")
            try:
                if locator.count() > 0 and locator.first.is_visible():
                    return locator.first
            except Exception:
                pass
        sleep(300)
    return None


def click_menu_button(page, text: str, ru_text: Optional[str] = None, max_retries: int = 4):
    candidates = []
    if text:
        candidates.append(text)
    if ru_text:
        candidates.append(ru_text)

    if not candidates:
        raise RuntimeError("菜单文本为空")

    for attempt in range(1, max_retries + 1):
        log(f"🎯 查找菜单：{text or ru_text}，第 {attempt}/{max_retries} 次")

        # 1. role button
        for name in candidates:
            try:
                loc = page.get_by_role("button", name=name)
                if loc.count() > 0 and loc.first.is_visible():
                    loc.first.scroll_into_view_if_needed()
                    sleep(300)
                    loc.first.click(timeout=5000)
                    sleep(2500)
                    return True
            except Exception:
                pass

        # 2. 文本点击
        for name in candidates:
            try:
                loc = page.get_by_text(name, exact=False)
                count = loc.count()
                for i in range(count):
                    item = loc.nth(i)
                    if item.is_visible():
                        item.scroll_into_view_if_needed()
                        sleep(300)
                        item.click(timeout=5000)
                        sleep(2500)
                        return True
            except Exception:
                pass

        # 3. XPath
        selectors = build_text_xpaths(text, ru_text)
        target = find_visible_by_xpaths(page, selectors, timeout_ms=3000)
        if target:
            try:
                target.scroll_into_view_if_needed()
                sleep(300)
                target.click(timeout=5000)
                sleep(2500)
                return True
            except Exception as e:
                log(f"⚠️ XPath 点击菜单失败：{text or ru_text}，原因：{e}")

        # 4. 尝试滚动
        try:
            page.mouse.wheel(0, 1000)
        except Exception:
            pass
        sleep(800)

    raise RuntimeError(f"找不到或无法点击菜单按钮: {text or ru_text}")


# =========================
# 页面类型识别
# =========================
def detect_page_type(page):
    """统一委托给页面状态识别器。"""
    return _detect_page_type(page)


def is_messenger_page(page) -> bool:
    return _is_messenger_page(page)


def is_company_select_page(page) -> bool:
    return detect_page_type(page) == "company_select"


def is_chat_detail_page(page) -> bool:
    try:
        url = page.url or ""
    except Exception:
        url = ""
    return "/app/messenger/" in url and ("group=" in url or "id=" in url)


def normalize_messenger_home(page):
    if is_chat_detail_page(page):
        log("ℹ️ 当前处于具体会话页，准备回到 messenger 首页")
        page.goto(TARGET_URL, wait_until="domcontentloaded", timeout=60000)
        sleep(4000)
        log(f"✅ 已回到 messenger 首页: {page.url}")


def wait_for_url_contains(page, keyword: str, timeout_ms: int = 20000) -> bool:
    deadline = time.time() + timeout_ms / 1000
    while time.time() < deadline:
        try:
            if keyword in page.url:
                return True
        except Exception:
            pass
        sleep(300)
    return False


# =========================
# 公司选择页处理
# =========================
def click_next_button(page, timeout_ms=10000):
    candidate_texts = ["下一步", "Далее", "Next"]

    deadline = time.time() + timeout_ms / 1000
    last_error = None

    while time.time() < deadline:
        for text in candidate_texts:
            try:
                btn = page.get_by_role("button", name=text)
                if btn.count() > 0 and btn.first.is_visible():
                    btn.first.scroll_into_view_if_needed()
                    sleep(300)
                    btn.first.click(timeout=5000)
                    log(f"✅ 已点击按钮: {text}")
                    return True
            except Exception as e:
                last_error = e

        for text in candidate_texts:
            try:
                btn = page.get_by_text(text, exact=True)
                if btn.count() > 0 and btn.first.is_visible():
                    btn.first.scroll_into_view_if_needed()
                    sleep(300)
                    btn.first.click(timeout=5000)
                    log(f"✅ 已通过文本点击按钮: {text}")
                    return True
            except Exception as e:
                last_error = e

        sleep(500)

    log(f"⚠️ 点击下一步失败: {last_error}")
    return False


def click_first_company_option(page):
    exclude_words = [
        "添加公司", "退出", "下一步",
        "Добавить компанию", "Выйти", "Далее",
        "帮助中心", "个人信息处理的条件",
        "中文", "Русский", "English",
    ]

    radio_selectors = [
        "[role='radio']",
        "label:has(input[type='radio'])",
        "input[type='radio']",
    ]

    for selector in radio_selectors:
        try:
            loc = page.locator(selector)
            count = loc.count()
            log(f"尝试公司候选选择器 {selector}，数量: {count}")

            for i in range(count):
                try:
                    item = loc.nth(i)
                    if not item.is_visible():
                        continue

                    clickable = item
                    if selector == "input[type='radio']":
                        try:
                            parent = item.locator("xpath=ancestor::*[self::label or self::div][1]")
                            if parent.count() > 0:
                                clickable = parent.first
                        except Exception:
                            clickable = item

                    clickable.scroll_into_view_if_needed()
                    sleep(300)
                    clickable.click(timeout=5000)
                    log(f"✅ 已点击公司候选项（radio类），第 {i + 1} 个")
                    sleep(1000)
                    return True
                except Exception as e:
                    log(f"⚠️ 点击 radio 候选失败，第 {i + 1} 个: {e}")
        except Exception as e:
            log(f"⚠️ 枚举选择器失败 {selector}: {e}")

    candidate_selectors = ["label", "div"]

    for selector in candidate_selectors:
        try:
            loc = page.locator(selector)
            count = loc.count()
            log(f"尝试公司卡片选择器 {selector}，数量: {count}")

            for i in range(count):
                try:
                    item = loc.nth(i)
                    if not item.is_visible():
                        continue

                    txt = item.inner_text(timeout=500).strip()
                    if not txt:
                        continue

                    if any(word in txt for word in exclude_words):
                        continue

                    looks_like_company = (
                        ("@" in txt) or
                        ("*" in txt) or
                        ("+" in txt) or
                        ("\n" in txt)
                    )
                    if not looks_like_company:
                        continue

                    if len(txt) > 300:
                        continue

                    item.scroll_into_view_if_needed()
                    sleep(300)
                    item.click(timeout=5000)
                    log(f"✅ 已点击公司候选项（文本结构识别），第 {i + 1} 个")
                    log(f"候选内容: {txt[:120]}")
                    sleep(1000)
                    return True
                except Exception as e:
                    log(f"⚠️ 点击公司卡片失败，第 {i + 1} 个: {e}")
        except Exception as e:
            log(f"⚠️ 枚举公司卡片失败 {selector}: {e}")

    return False


def handle_company_select(page):
    log("检测到公司选择页，开始处理...")
    dump_page_state(page, "company_select_before")

    selected = click_first_company_option(page)
    if not selected:
        log("⚠️ 未能识别并点击公司项")
        dump_page_state(page, "error_company_option_not_found")
        return False

    next_ok = click_next_button(page, timeout_ms=10000)
    if not next_ok:
        log("⚠️ 未能点击下一步")
        dump_page_state(page, "error_click_next_after_company_select")
        return False

    try:
        page.wait_for_load_state("networkidle", timeout=15000)
    except Exception:
        pass

    sleep(3000)
    dump_page_state(page, "company_select_after")
    return True


# =========================
# OTP 登录恢复
# =========================
def is_on_otp_page(page):
    body_text = safe_body_text(page)
    try:
        current_url = page.url
    except Exception:
        current_url = ""

    return (
        "Введите код" in body_text
        or "Отправили код на почту" in body_text
        or "Получить новый код" in body_text
        or "输入验证码" in body_text
        or "/otp" in current_url
        or "id.ozon.ru/otp" in current_url
    )


def log_input_candidates(page, tag="OTP_INPUT_DEBUG"):
    try:
        inputs = page.locator("input")
        count = inputs.count()
        log(f"[{tag}] 当前页面 input 数量: {count}")

        for i in range(count):
            try:
                el = inputs.nth(i)
                info = el.evaluate(
                    """e => ({
                        type: e.getAttribute('type'),
                        name: e.getAttribute('name'),
                        value: e.value || '',
                        placeholder: e.getAttribute('placeholder'),
                        className: e.className || '',
                        inputMode: e.getAttribute('inputmode'),
                        autocomplete: e.getAttribute('autocomplete'),
                        outerHTML: e.outerHTML.slice(0, 300)
                    })"""
                )
                log(f"[{tag}] input[{i}]: {info}")
            except Exception as e:
                log(f"[{tag}] 读取 input[{i}] 失败: {e}")
    except Exception as e:
        log(f"[{tag}] 枚举 input 失败: {e}")


def get_otp_input_locator(page):
    if not is_on_otp_page(page):
        return None, None

    try:
        named = page.locator("input[name='otp']")
        if named.count() > 0:
            first = named.first
            if first.is_visible():
                return "otp_named", first
    except Exception:
        pass

    try:
        text_inputs = page.locator("input[type='text']")
        count = text_inputs.count()

        visible_candidates = []
        for i in range(count):
            try:
                el = text_inputs.nth(i)
                if not el.is_visible():
                    continue

                    # no-op
                info = el.evaluate(
                    """e => ({
                        value: e.value || '',
                        className: e.className || '',
                        placeholder: e.getAttribute('placeholder') || '',
                        inputMode: e.getAttribute('inputmode') || '',
                        autocomplete: e.getAttribute('autocomplete') || '',
                        outerHTML: e.outerHTML.slice(0, 300)
                    })"""
                )
                visible_candidates.append((el, info))
            except Exception:
                continue

        for el, info in visible_candidates:
            value = info.get("value", "")
            class_name = info.get("className", "")
            outer_html = info.get("outerHTML", "")

            if "−" in value or "—" in value:
                return "otp_masked_text", el

            if "input-code" in class_name.lower() or "dsinputcode" in outer_html.lower():
                return "otp_masked_text", el

        if len(visible_candidates) == 1:
            return "otp_generic_text", visible_candidates[0][0]

        for el, info in visible_candidates:
            value = (info.get("value", "") or "").replace(" ", "")
            if len(value) <= 10:
                return "otp_generic_text", el

    except Exception:
        pass

    return None, None


def wait_for_otp_input(page, timeout_ms=30000):
    start = time.time()
    timeout_sec = timeout_ms / 1000

    while time.time() - start < timeout_sec:
        kind, locator = get_otp_input_locator(page)
        if kind and locator:
            try:
                locator.wait_for(state="visible", timeout=1000)
                return kind, locator
            except Exception:
                pass

        page_type = detect_page_type(page)
        if page_type in ("blocked", "dashboard", "messenger", "company_select"):
            return "page_changed", None

        sleep(500)

    return None, None


def clear_and_type_otp(otp_input, otp_code, page, input_kind):
    otp_input.wait_for(state="visible", timeout=10000)
    otp_input.click()
    sleep(300)

    if input_kind == "otp_named":
        try:
            otp_input.fill("")
        except Exception:
            try:
                otp_input.press("Control+A")
                otp_input.press("Backspace")
            except Exception:
                pass

        otp_input.type(otp_code, delay=180)
        return

    try:
        otp_input.press("Control+A")
        otp_input.press("Backspace")
        sleep(200)
    except Exception:
        pass

    for ch in otp_code:
        otp_input.type(ch, delay=180)
        sleep(80)


def wait_for_post_otp_result(page, timeout_ms=20000):
    start = time.time()
    timeout_sec = timeout_ms / 1000

    while time.time() - start < timeout_sec:
        page_type = detect_page_type(page)

        if page_type in ("dashboard", "company_select", "messenger", "blocked"):
            return page_type

        try:
            body_text = page.locator("body").inner_text(timeout=1000)

            if "Произошла ошибка. Попробуйте еще раз" in body_text:
                return "otp_error"
            if "Неверный код" in body_text or "Неправильный код" in body_text:
                return "otp_error"
            if "验证码错误" in body_text or "验证码无效" in body_text:
                return "otp_error"
        except Exception:
            pass

        sleep(500)

    final_type = detect_page_type(page)
    if final_type == "otp":
        return "otp_still_here"
    return final_type


def login_with_email_otp(page, context, account: OzonAccount):
    log(f"开始邮箱验证码登录流程: {account.email}")

    current_type = detect_page_type(page)
    log(f"登录流程起始页面类型: {current_type}")

    if current_type == "messenger":
        log("✅ 已在 messenger 页面，无需登录")
        return True

    if current_type == "company_select":
        ok = handle_company_select(page)
        if ok:
            save_login_state(context, account.storage_path)
        return ok

    if current_type == "login":
        try:
            dump_page_state(page, "before_click_login")

            # 支持点击中文或俄文的登录按钮
            login_btn = None
            try:
                # 先尝试查找俄文按钮 "Войти"
                login_btn = page.get_by_text("Войти", exact=True).first
                login_btn.wait_for(state="visible", timeout=5000)
            except:
                # 如果俄文按钮未找到，尝试查找中文按钮 "登录"
                login_btn = page.get_by_text("登录", exact=True).first
                login_btn.wait_for(state="visible", timeout=5000)

            login_btn.click()
            page.wait_for_load_state("networkidle", timeout=20000)
            log("✅ 已点击登录按钮")

            dump_page_state(page, "after_click_login")
        except Exception as e:
            log(f"⚠️ 点击登录按钮失败: {e}")
            dump_page_state(page, "error_click_login")
            return False

    current_type = detect_page_type(page)
    log(f"点击登录后页面类型: {current_type}")

    if current_type == "ozon_id_phone":
        try:
            dump_page_state(page, "before_click_email_login")

            # 支持点击中文或俄文的邮箱登录按钮
            email_login_btn = None
            try:
                # 先尝试查找俄文按钮 "Войти по почте"
                email_login_btn = page.get_by_text("Войти по почте", exact=True).first
                email_login_btn.wait_for(state="visible", timeout=5000)
            except:
                # 如果俄文按钮未找到，尝试查找中文按钮 "邮箱登录"
                email_login_btn = page.get_by_text("使用邮箱登录", exact=True).first
                email_login_btn.wait_for(state="visible", timeout=5000)

            email_login_btn.click()
            page.wait_for_load_state("networkidle", timeout=15000)
            log("✅ 已点击“通过邮箱登录”")

            dump_page_state(page, "after_click_email_login")
        except Exception as e:
            log(f"⚠️ 点击“通过邮箱登录”失败: {e}")
            dump_page_state(page, "error_click_email_login")
            return False
    elif current_type not in ("otp",):
        log("⚠️ 当前不是预期的 Ozon ID 手机号登录页")
        dump_page_state(page, "error_unexpected_login_stage")
        return False

    try:
        if detect_page_type(page) != "otp":
            page.wait_for_selector("input[type='email']", timeout=20000)
            log("✅ 邮箱输入框已出现")
            dump_page_state(page, "email_input_visible")
    except Exception as e:
        log(f"⚠️ 未找到邮箱输入框: {e}")
        dump_page_state(page, "error_email_input_not_found")
        return False

    if detect_page_type(page) != "otp":
        baseline_mail_id = get_latest_ozon_mail_id()
        log(f"提交前基线邮件 ID: {baseline_mail_id}")

        try:
            page.fill("input[type='email']", account.email)
            request_time = datetime.now(timezone.utc)

            submit_btn = page.locator("button[type='submit']").first
            submit_btn.click()
            log("✅ 已提交邮箱")
            log(f"验证码请求时间(UTC): {request_time.isoformat()}")

            dump_page_state(page, "after_submit_email")
        except Exception as e:
            log(f"⚠️ 填写邮箱或提交失败: {e}")
            dump_page_state(page, "error_submit_email")
            return False
    else:
        baseline_mail_id = get_latest_ozon_mail_id()
        request_time = datetime.now(timezone.utc)

    try:
        log("已提交邮箱，等待验证码输入框...")

        otp_kind, otp_input = wait_for_otp_input(page, timeout_ms=30000)

        if otp_kind == "page_changed":
            log("⚠️ 等待验证码输入框时页面已跳转")
            dump_page_state(page, "error_otp_page_changed_before_input")
            return False

        if not otp_input:
            log("⚠️ 未识别到验证码输入框")
            log_input_candidates(page, "OTP_INPUT_NOT_FOUND")
            dump_page_state(page, "error_otp_input_not_found")
            return False

        log(f"✅ 验证码输入框已出现，识别类型: {otp_kind}")
        log_input_candidates(page, "OTP_INPUT_FOUND")
        dump_page_state(page, "otp_input_visible")

    except Exception as e:
        log(f"⚠️ 未出现验证码输入框: {e}")
        log_input_candidates(page, "OTP_INPUT_WAIT_EXCEPTION")
        dump_page_state(page, "error_otp_input_not_found")
        return False

    # 强制采用手动输入验证码（不再尝试 IMAP 自动提取）
    import tkinter as tk
    from tkinter import simpledialog

    root = tk.Tk()
    root.withdraw()
    otp = simpledialog.askstring("验证码输入", "请输入收到的验证码:", parent=root)
    root.destroy()

    if not otp:
        log("❌ 用户未输入验证码")
        return False

    log("✅ 已接收手动输入验证码")

    try:
        log(f"正在填入验证码: {otp}")

        otp_kind, otp_input = get_otp_input_locator(page)
        if not otp_input:
            log("⚠️ 填写验证码前无法重新定位输入框")
            log_input_candidates(page, "OTP_INPUT_RELOCATE_FAILED")
            dump_page_state(page, "error_fill_otp_locator_missing")
            return False

        clear_and_type_otp(otp_input, otp, page, otp_kind)

        try:
            current_value = otp_input.input_value()
            log(f"当前输入框中的验证码值: {current_value}")
        except Exception:
            log("⚠️ 输入后页面可能已跳转，无法再读取 otp input 值")

        dump_page_state(page, "after_fill_otp")

        log("验证码已填写，等待页面自动跳转...")
        sleep(5000)
        dump_page_state(page, "after_fill_otp_wait")

    except Exception as e:
        log(f"⚠️ 填写验证码失败: {e}")
        log_input_candidates(page, "OTP_INPUT_FILL_EXCEPTION")
        dump_page_state(page, "error_fill_otp")
        return False

    result_type = wait_for_post_otp_result(page, timeout_ms=20000)
    log(f"OTP 提交后的页面类型: {result_type}")

    dump_page_state(page, "after_otp_result")

    if result_type == "otp_error":
        log("❌ 页面明确提示验证码提交失败")
        return False

    if result_type == "otp_still_here":
        log("⚠️ 输入验证码后仍停留在 OTP 页面")
        return False

    if result_type == "blocked":
        log("❌ OTP 成功后被风控拦截")
        return False

    if result_type == "company_select":
        ok = handle_company_select(page)
        if not ok:
            return False

        try:
            page.goto(TARGET_URL, wait_until="domcontentloaded", timeout=60000)
            sleep(3000)
        except Exception:
            pass

        save_login_state(context, account.storage_path)
        return True

    if result_type in ("dashboard", "messenger"):
        save_login_state(context, account.storage_path)
        return True

    log("⚠️ 验证码提交后页面状态未知")
    return False


# =========================
# 登录保障入口
# =========================
def ensure_logged_in_and_ready(page, context, account: OzonAccount):
    log(f"🌐 登录后检查页面: {page.url}")
    sleep(3000)

    page_type = detect_page_type(page)
    log(f"当前页面类型: {page_type}")

    if is_messenger_page(page):
        log("✅ 已进入 messenger 页面")
        return

    if page_type in ("login", "ozon_id_phone", "otp", "company_select"):
        log("ℹ️ 检测到登录态失效或登录流程页面，开始自动恢复登录")
        ok = login_with_email_otp(page, context, account)
        if not ok:
            raise RuntimeError("自动登录恢复失败")

        try:
            page.goto(TARGET_URL, wait_until="domcontentloaded", timeout=60000)
            sleep(3000)
        except Exception:
            pass

        if is_messenger_page(page):
            save_login_state(context, account.storage_path)
            log("✅ 自动登录恢复后已进入 messenger 页面")
            return

    if is_company_select_page(page):
        log("ℹ️ 检测到公司选择页，准备选择公司并点击下一步")

        ok = handle_company_select(page)
        if not ok:
            raise RuntimeError("公司选择页处理失败")

        if wait_for_url_contains(page, "/app/messenger", timeout_ms=15000):
            save_login_state(context, account.storage_path)
            log("✅ 点击下一步后已进入 messenger 页面")
            return

        log("⚠️ 点击下一步后未自动进入目标页，尝试手动打开 TARGET_URL")
        page.goto(TARGET_URL, wait_until="domcontentloaded", timeout=60000)
        sleep(3000)

        if is_messenger_page(page):
            save_login_state(context, account.storage_path)
            log("✅ 手动跳转后已进入 messenger 页面")
            return

    if "/signin" in page.url or "/registration" in page.url:
        log("⚠️ 当前仍在登录相关页面，尝试再次打开目标页")
        page.goto(TARGET_URL, wait_until="domcontentloaded", timeout=60000)
        sleep(3000)

        if is_messenger_page(page):
            save_login_state(context, account.storage_path)
            log("✅ 重试后已进入 messenger 页面")
            return

        page_type = detect_page_type(page)
        if page_type in ("login", "ozon_id_phone", "otp", "company_select"):
            ok = login_with_email_otp(page, context, account)
            if not ok:
                raise RuntimeError("再次尝试自动登录恢复失败")

            page.goto(TARGET_URL, wait_until="domcontentloaded", timeout=60000)
            sleep(3000)
            if is_messenger_page(page):
                save_login_state(context, account.storage_path)
                log("✅ 再次自动登录恢复后已进入 messenger 页面")
                return

    if is_messenger_page(page):
        save_login_state(context, account.storage_path)
        return

    raise RuntimeError(f"登录后未能进入目标页面，当前 URL: {page.url}")


# =========================
# 业务页面查找
# =========================
def find_sku_input(page, max_retries: int = 5):
    for attempt in range(1, max_retries + 1):
        locator = page.locator("textarea")
        try:
            count = locator.count()
            for i in range(count):
                item = locator.nth(i)
                if item.is_visible():
                    log(f"✅ 找到 SKU 输入框 textarea，第 {i + 1} 个")
                    return item
        except Exception:
            pass

        try:
            locator = page.locator("input[type='text']")
            count = locator.count()
            for i in range(count):
                item = locator.nth(i)
                if item.is_visible():
                    log(f"✅ 找到 SKU 输入框 input[type='text']，第 {i + 1} 个")
                    return item
        except Exception:
            pass

        log(f"⏳ 查找 SKU 输入框，第 {attempt}/{max_retries} 次")
        sleep(1000)

    raise RuntimeError("未找到 SKU 输入框 textarea / input[type='text']")


def set_input_value(locator, value: str):
    locator.click()
    sleep(200)
    locator.fill("")
    sleep(150)
    locator.fill(value)


def press_enter(locator):
    locator.press("Enter")


def wait_for_search_result(page, sku: str, timeout_ms: int = 10000):
    deadline = time.time() + timeout_ms / 1000
    while time.time() < deadline:
        try:
            file_input = page.locator("input[type='file']")
            if file_input.count() > 0:
                return True
        except Exception:
            pass

        try:
            body_text = page.locator("body").inner_text(timeout=800)
            if sku in body_text:
                return True
        except Exception:
            pass

        sleep(500)

    raise RuntimeError(f"等待 SKU 查询结果超时: {sku}")


def find_file_input(page, max_retries: int = 5):
    for attempt in range(1, max_retries + 1):
        selectors = [
            "input[type='file'][accept*='image']",
            "input[type='file']",
        ]

        for selector in selectors:
            locator = page.locator(selector)
            try:
                count = locator.count()
                if count > 0:
                    for i in range(count - 1, -1, -1):
                        chosen = locator.nth(i)
                        try:
                            if chosen.is_visible():
                                log(f"✅ 找到上传控件：{selector}，第 {i + 1} 个")
                                return chosen
                        except Exception:
                            pass

                    chosen = locator.nth(count - 1)
                    log(f"✅ 找到上传控件：{selector}，第 {count} 个（不可见兜底）")
                    return chosen
            except Exception:
                pass

        log(f"⏳ 查找上传控件，第 {attempt}/{max_retries} 次")
        sleep(1000)

    raise RuntimeError("未找到图片上传 input[type='file']")


def get_send_button_from_file_input(page, file_input):
    candidates = []

    try:
        locator = page.locator("input[type='file'] + div + button")
        if locator.count() > 0:
            candidates.append(locator.last)
    except Exception:
        pass

    try:
        locator = page.locator("input[type='file'] ~ button")
        if locator.count() > 0:
            candidates.append(locator.last)
    except Exception:
        pass

    try:
        locator = file_input.locator("xpath=following-sibling::button[1]")
        if locator.count() > 0:
            candidates.append(locator.first)
    except Exception:
        pass

    try:
        locator = file_input.locator("xpath=following-sibling::*//button[1]")
        if locator.count() > 0:
            candidates.append(locator.first)
    except Exception:
        pass

    try:
        for name in ["发送", "Отправить", "Send"]:
            locator = page.get_by_role("button", name=name)
            if locator.count() > 0:
                candidates.append(locator.last)
    except Exception:
        pass

    for idx, btn in enumerate(candidates, 1):
        try:
            if btn.count() > 0 and btn.first.is_visible():
                log(f"✅ 发送按钮定位成功，候选 #{idx}")
                return btn.first
        except Exception:
            pass

    try:
        locator = page.locator("xpath=(//input[@type='file']/following-sibling::button)[last()]")
        if locator.count() > 0 and locator.first.is_visible():
            log("✅ 发送按钮定位成功，使用全局 XPath 兜底")
            return locator.first
    except Exception:
        pass

    raise RuntimeError("未找到发送按钮")


def wait_for_send_button_enabled(page, file_input, timeout_ms: int = 15000):
    deadline = time.time() + timeout_ms / 1000
    last_state = ""

    while time.time() < deadline:
        try:
            btn = get_send_button_from_file_input(page, file_input)

            visible = btn.is_visible()
            disabled = btn.is_disabled()
            enabled = not disabled

            last_state = f"visible={visible}, disabled={disabled}, enabled={enabled}"
            log(f"⏳ 发送按钮状态: {last_state}")

            if visible and enabled:
                return btn
        except Exception as e:
            last_state = str(e)

        sleep(300)

    raise RuntimeError(f"等待发送按钮可用超时，最后状态: {last_state}")


def click_send_button(page, file_input, timeout_ms: int = 15000):
    btn = wait_for_send_button_enabled(page, file_input, timeout_ms=timeout_ms)

    try:
        btn.scroll_into_view_if_needed()
    except Exception:
        pass

    try:
        btn.click(timeout=5000)
        log("✅ 已点击发送按钮")
        return
    except Exception as e:
        log(f"⚠️ 普通点击发送按钮失败：{e}")

    try:
        btn.click(force=True, timeout=5000)
        log("✅ 已强制点击发送按钮")
        return
    except Exception as e:
        log(f"⚠️ 强制点击发送按钮失败：{e}")

    raise RuntimeError("发送按钮点击失败")


def wait_for_upload_finished(page, timeout_ms: int = 20000):
    success_texts = ["上传成功", "已上传", "上传完成", "успешно", "загружено", "готово", "success"]
    error_texts = ["上传失败", "失败", "error", "ошибка"]

    deadline = time.time() + timeout_ms / 1000
    while time.time() < deadline:
        try:
            body_text = page.locator("body").inner_text(timeout=1000).lower()

            for t in error_texts:
                if t.lower() in body_text:
                    raise RuntimeError("检测到上传失败提示")

            for t in success_texts:
                if t.lower() in body_text:
                    return True
        except PlaywrightTimeoutError:
            pass

        sleep(800)

    return True


# =========================
# 业务流程
# =========================
def navigate_menu(page):
    normalize_messenger_home(page)

    menu_buttons = [
        {"text": "商品和价格", "ru_text": "Товары и цены"},
        {"text": "质量监督", "ru_text": "Контроль качества"},
        {"text": "卖家使用我的品牌", "ru_text": "Продавцы используют мой бренд"},
    ]

    for idx, item in enumerate(menu_buttons, 1):
        text = (item.get("text") or "").strip()
        ru_text = (item.get("ru_text") or "").strip()

        if not text and not ru_text:
            log(f"⚠️ 菜单配置为空，跳过第 {idx} 项")
            continue

        log(f"🎯 菜单导航 {idx}/{len(menu_buttons)}: {text or ru_text}")
        click_menu_button(page, text or ru_text, ru_text or None)


def process_single_sku(page, sku: str, image_path: str):
    log(f"📦 开始处理 SKU: {sku}")

    sku_input = find_sku_input(page)
    set_input_value(sku_input, sku)

    press_enter(sku_input)
    sleep(2000)

    wait_for_search_result(page, sku)

    file_input = find_file_input(page)
    file_input.set_input_files(image_path)
    log("✅ 图片已选择，等待发送按钮")
    sleep(1500)

    click_send_button(page, file_input, timeout_ms=20000)
    sleep(2000)

    wait_for_upload_finished(page)
    log(f"✅ SKU 处理完成: {sku}")


def execute(page, skus: List[str], image_path: str):
    navigate_menu(page)
    log(f"✅ 菜单导航完成，开始处理 {len(skus)} 个 SKU")

    for i, sku in enumerate(skus, 1):
        log(f"➡️ {i}/{len(skus)}")
        process_single_sku(page, sku, image_path)
        sleep(1200)


# =========================
# 对外入口
# =========================
def prepare_browser(
    email: str,
    imap_password: str = "",
    storage_path: str = None,
    headless: bool = False,
    slow_mo: int = 200,
    use_manual_login: bool = False,
):
    """准备浏览器 - 启动浏览器并确保登录状态"""
    if not session_service:
        raise RuntimeError("未初始化会话服务，请先调用 set_logger")

    account = OzonAccount(email, imap_password, storage_path, use_manual_login)
    session = session_service.get_session(account, headless, slow_mo)

    session.page.goto(TARGET_URL, wait_until="domcontentloaded", timeout=60000)
    log(f"当前页面: {session.page.url}")
    log(f"当前标题: {session.page.title()}")

    if not page_service:
        raise RuntimeError("未初始化页面服务，请先调用 set_logger")

    page_service.ensure_logged_in_and_ready(session.page, session.context, account, TARGET_URL)

    try:
        from services.utils import save_login_state
        save_login_state(session.context, account.storage_path)
        log(f"✅ 已刷新并保存登录态: {account.storage_path}")
    except Exception as e:
        log(f"⚠️ 保存登录态失败: {e}")

    page_service.normalize_messenger_home(session.page, TARGET_URL)

    log(f"浏览器准备完成，当前页面: {session.page.url}")
    return session.page


def run_task(
    email: str,
    excel_path: str,
    image_path: str,
    imap_password: str,
    storage_path: str = None,
    headless: bool = False,
    slow_mo: int = 200,
):
    """兼容入口：从 Excel 读取 SKU 后执行任务"""
    skus = read_skus_from_excel(excel_path)
    return run_task_with_skus(
        email=email,
        skus=skus,
        image_path=image_path,
        imap_password=imap_password,
        storage_path=storage_path,
        headless=headless,
        slow_mo=slow_mo,
    )


def run_task_with_skus(
    email: str,
    skus: List[str],
    image_path: str,
    imap_password: str = "",
    storage_path: str = None,
    headless: bool = False,
    slow_mo: int = 200,
    use_manual_login: bool = False,
):
    """执行任务 - 直接使用 SKU 列表发送图片"""
    from services.utils import ensure_dirs
    ensure_dirs()

    if not os.path.exists(image_path):
        raise FileNotFoundError(f"未找到图片文件: {image_path}")

    normalized_skus = [str(s).strip() for s in skus if str(s).strip()]
    normalized_skus = list(dict.fromkeys(normalized_skus))

    if not normalized_skus:
        raise ValueError("未提供有效 SKU")

    log(f"📊 读取到 {len(normalized_skus)} 个 SKU")

    account = OzonAccount(email, imap_password, storage_path, use_manual_login)
    session = session_service.get_session(account, headless, slow_mo)

    try:
        # 检查会话是否有效
        from services.session_service import SessionService
        if not isinstance(session_service, SessionService) or not session_service._is_session_alive(session):
            raise RuntimeError("浏览器页面不可用")

        # 只有当前不在可用页面时，才重新跳目标页
        current_type = page_service.detect_page_type(session.page)
        log(f"执行前页面类型: {current_type}")

        if current_type not in ("messenger", "company_select", "login", "ozon_id_phone", "otp"):
            try:
                session.page.goto(TARGET_URL, wait_until="domcontentloaded", timeout=60000)
                sleep(3000)
            except Exception:
                pass

        page_service.ensure_logged_in_and_ready(session.page, session.context, account, TARGET_URL)

        try:
            from services.utils import save_login_state
            save_login_state(session.context, account.storage_path)
            log(f"✅ 已刷新并保存登录态: {account.storage_path}")
        except Exception as e:
            log(f"⚠️ 保存登录态失败: {e}")

        log(f"登录完成后页面: {session.page.url}")
        log(f"登录完成后标题: {session.page.title()}")

        from services.sku_service import MENU_BUTTONS
        summary = sku_service.execute(session.page, normalized_skus, image_path, MENU_BUTTONS)

        try:
            from services.utils import save_login_state
            save_login_state(session.context, account.storage_path)
            log(f"✅ 任务结束后已保存登录态: {account.storage_path}")
        except Exception as e:
            log(f"⚠️ 任务结束后保存登录态失败: {e}")

        log("✅ 当前批次任务执行完成，浏览器保持打开以便复用")
        return summary

    except Exception as e:
        log(f"❌ 错误: {e}")
        raise


def close_all_sessions():
    """关闭所有浏览器会话"""
    if session_service:
        session_service.close_all_sessions()


def close_session(email: str):
    """关闭指定邮箱的浏览器会话"""
    if session_service:
        session_service.close_session(email)


# =========================
# 示例 main
# =========================
def main():
    """
    默认演示：
    1. 启动并准备浏览器
    2. 执行一次任务
    3. 不关闭浏览器

    如果你想最后关闭浏览器，请手动调用：
        close_all_sessions()
    """

    # 使用示例邮箱
    test_email = "xpw709@163.com"

    prepare_browser(
        email=test_email,
        storage_path=None,  # 会自动生成
        headless=False,
        slow_mo=200,
    )

    run_task(
        email=test_email,
        excel_path="sku.xlsx",
        image_path="icon.png",
        storage_path=None,
        headless=False,
        slow_mo=200,
    )

    log("ℹ️ main() 结束，但浏览器不会关闭。")


if __name__ == "__main__":
    main()
